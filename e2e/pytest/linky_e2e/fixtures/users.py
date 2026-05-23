from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from linky_e2e.config import settings
from linky_e2e.storage.state import normalize_storage_state_path

TEST_USER_KEYS = (
    "user1",
    "user2",
    "user3",
    "user4",
    "user5",
    "user6",
    "user7",
)


@dataclass(frozen=True)
class TestUser:
    id: str
    email: str
    password: str
    otp: str
    storage_state_path: str
    first_name: str | None = None
    last_name: str | None = None


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _load_sheet_matrix(file_name: str) -> list[list[object | None]]:
    file_path = settings.test_data_dir / file_name
    if not file_path.exists():
        raise FileNotFoundError(f"Missing test data file: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def read_clerk_test_user_rows() -> list[TestUser]:
    matrix = _load_sheet_matrix("data_test_users.xlsx")
    out: list[TestUser] = []
    for i in range(1, len(matrix)):
        cols = matrix[i]
        if not cols or cols[0] is None or _cell_str(cols[0]) == "":
            continue
        out.append(
            TestUser(
                id=_cell_str(cols[0]),
                first_name=_cell_str(cols[1]) or None,
                last_name=_cell_str(cols[2]) or None,
                email=_cell_str(cols[3]),
                password=_cell_str(cols[4]),
                otp=_cell_str(cols[5]),
                storage_state_path=normalize_storage_state_path(_cell_str(cols[6])),
            )
        )
    return out


def build_test_users_registry() -> dict[str, TestUser]:
    rows = read_clerk_test_user_rows()
    by_id = {r.id: r for r in rows}
    out: dict[str, TestUser] = {}
    for key in TEST_USER_KEYS:
        u = by_id.get(key)
        if u is None:
            raise RuntimeError(f'Missing row with id "{key}" in data_test_users.xlsx')
        out[key] = u
    return out


_TEST_USERS_CACHE: dict[str, TestUser] | None = None


def get_test_users() -> dict[str, TestUser]:
    global _TEST_USERS_CACHE
    if _TEST_USERS_CACHE is None:
        _TEST_USERS_CACHE = build_test_users_registry()
    return _TEST_USERS_CACHE


class _TestUsersProxy:
    def __getitem__(self, key: str) -> TestUser:
        return get_test_users()[key]

    def get(self, key: str, default: TestUser | None = None) -> TestUser | None:
        return get_test_users().get(key, default)

    def __contains__(self, key: object) -> bool:
        return key in get_test_users()

    def keys(self):
        return get_test_users().keys()

    def values(self):
        return get_test_users().values()

    def items(self):
        return get_test_users().items()


TEST_USERS: _TestUsersProxy = _TestUsersProxy()


@dataclass
class LoginTestRow:
    sheet_row_index: int
    email: str
    password: str
    otp: str
    message: str


@dataclass
class SignupTestRow:
    sheet_row_index: int
    first_name: str
    last_name: str
    email: str
    password: str
    otp: str
    message: str


def read_login_test_rows() -> list[LoginTestRow]:
    matrix = _load_sheet_matrix("data_test_login.xlsx")
    out: list[LoginTestRow] = []
    for sheet_row in range(2, 14):
        cols = matrix[sheet_row - 1] if sheet_row - 1 < len(matrix) else None
        if not cols or all(c is None for c in cols[:3]):
            continue
        out.append(
            LoginTestRow(
                sheet_row_index=sheet_row,
                email=_cell_str(cols[0]),
                password=_cell_str(cols[1]),
                otp=_cell_str(cols[2]),
                message=_cell_str(cols[3]) if len(cols) > 3 else "",
            )
        )
    return out


def read_signup_test_rows() -> list[SignupTestRow]:
    matrix = _load_sheet_matrix("data_test_signup.xlsx")
    out: list[SignupTestRow] = []
    for sheet_row in range(2, 36):
        cols = matrix[sheet_row - 1] if sheet_row - 1 < len(matrix) else None
        if not cols or all(c is None for c in cols[:5]):
            continue
        out.append(
            SignupTestRow(
                sheet_row_index=sheet_row,
                first_name=_cell_str(cols[0]),
                last_name=_cell_str(cols[1]),
                email=_cell_str(cols[2]),
                password=_cell_str(cols[3]),
                otp=_cell_str(cols[4]),
                message=_cell_str(cols[5]) if len(cols) > 5 else "",
            )
        )
    return out
